
from PySide6.QtWidgets import QMainWindow, QApplication, QMessageBox
from PySide6.QtGui import QPixmap, QImage
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import sys
import os
import openpyxl
import math
from openpyxl import Workbook
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, Border, Side
from ui_Test1 import Ui_MainWindow
from ui_Test2 import Ui_MainWindow2
from ui_Test3 import Ui_MainWindow3
from ui_Test4 import Ui_MainWindow4

# Define global variables for material properties and constants
E = 210e9  # Young's modulus (Pa)
A = 0.01  # Cross-sectional area (m^2)
density = 7850  # kg/m^3
gravity = 9.81  # m/s^2
material_elasticity_kg_per_mm2 = 345e6  # 항복응력 (N/m^2), 강재의 일반적인 값

def load_image(image_file, label):
    resources_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
    image_path = os.path.join(resources_dir, image_file)
    pixmap = QPixmap(image_path)
    label.setPixmap(pixmap)

class BridgeElement:
    def __init__(self, material_type, bridge_length_m, support_points_count, material_elasticity_kg_per_mm2,
                 live_load_kN, member_section):
        self.material_type = material_type  # 재료 종류
        self.bridge_length_m = bridge_length_m  # 교량 길이 (m)
        self.support_points_count = support_points_count  # 절점 갯수 (개)
        self.material_elasticity_kg_per_mm2 = material_elasticity_kg_per_mm2  # 재료 탄성 계수 (kg/mm^2)
        self.live_load_kN = live_load_kN  # 활 하중 (kN)
        self.member_section = member_section  # 부재 단면

    def set_fixed_load(self, fixed_load_kN):
        self.fixed_load_kN = fixed_load_kN

    def set_unit_weight(self, unit_weight_kg):
        self.unit_weight_kg = unit_weight_kg

def calculate_weight(member_section, bridge_length_m, N):
    # 자중 계산 함수
    unit_weight_kg = density * gravity
    bottom_members_count = N // 2
    volume = member_section * bridge_length_m / bottom_members_count
    weight = volume * unit_weight_kg
    fixed_load_kN = weight / bridge_length_m

    return weight, fixed_load_kN

def calculate_section_modulus(member_section):
    width = (member_section ** 0.5) / 2  # 가정: 플랜지와 웹의 비율에 따른 폭
    height = member_section / width  # 단면적에 따른 높이
    S = (width * height ** 2) / 6  # 대략적인 단면적 모멘트 계산

    return S

def calculate_flexural_strength(member_section, bridge_length_m, live_load_kN, fixed_load_kN):
    D = fixed_load_kN * bridge_length_m ** 2 / 8
    L = live_load_kN * bridge_length_m / 4
    Mu = 1.2 * D + 1.6 * L
    return Mu, D, L

def evaluate_safety(Mu, Mn):
    return "안전" if Mu <= Mn else "불안전"


# 트러스 구조 생성 함수 수정
def create_truss_structure(n):
    nodes = []
    elements = []

    # 노드 생성
    for i in range(n):
        nodes.append([i, 0])
        nodes.append([i, 1])

    # 요소 생성
    for i in range(n):
        if i < n - 1:
            elements.append([2 * i, 2 * (i + 1)])      # 아래 가로선
            elements.append([2 * i + 1, 2 * (i + 1) + 1])  # 위 가로선
            elements.append([2 * i, 2 * (i + 1) + 1])  # 대각선 (아래에서 위로)
            elements.append([2 * i + 1, 2 * (i + 1)])  # 대각선 (위에서 아래로)

    # 세로선
    for i in range(n):
        elements.append([2 * i, 2 * i + 1])

    nodes = np.array(nodes)
    elements = np.array(elements)
    return nodes, elements

# 요소의 길이와 각도 계산
def element_properties(node1, node2):
    length = np.linalg.norm(node2 - node1)
    angle = np.arctan2(node2[1] - node1[1], node2[0] - node1[0])
    return length, angle

# 요소 강성 행렬 계산
def element_stiffness_matrix(length, angle):
    c = np.cos(angle)
    s = np.sin(angle)
    k = (E * A / length) * np.array([
        [c * c, c * s, -c * c, -c * s],
        [c * s, s * s, -c * s, -s * s],
        [-c * c, -c * s, c * c, c * s],
        [-c * s, -s * s, c * s, s * s]
    ])
    return k

# 전체 강성 행렬 조립
def assemble_global_stiffness(nodes, elements):
    n_nodes = nodes.shape[0]
    n_dof = 2 * n_nodes
    K = np.zeros((n_dof, n_dof))

    for element in elements:
        if element[0] < n_nodes and element[1] < n_nodes:
            node1 = nodes[element[0]]
            node2 = nodes[element[1]]
            length, angle = element_properties(node1, node2)
            k = element_stiffness_matrix(length, angle)

            dof_indices = np.array([
                2 * element[0], 2 * element[0] + 1,
                2 * element[1], 2 * element[1] + 1
            ])

            for i in range(4):
                for j in range(4):
                    K[dof_indices[i], dof_indices[j]] += k[i, j]
        else:
            print(f"Invalid element: {element}")

    return K

# 응력 계산
def element_stress(node1, node2, u1, u2, length, angle):
    c = np.cos(angle)
    s = np.sin(angle)
    T = np.array([-c, -s, c, s])
    strain = (1 / length) * T @ np.concatenate((u1, u2))
    stress = E * strain
    return stress


# 시각화 함수
def plot_truss(nodes, elements):
    plt.figure(figsize=(8.2, 5))

    for idx, element in enumerate(elements):
        node1 = nodes[element[0]]
        node2 = nodes[element[1]]
        x_values = [node1[0], node2[0]]
        y_values = [node1[1], node2[1]]
        plt.plot(x_values, y_values, 'b-o')

        # Display element number slightly off-center
        offset_x = (node2[0] - node1[0]) * 0.05
        offset_y = (node2[1] - node1[1]) * 0.05
        plt.text(np.mean(x_values) + offset_x, np.mean(y_values) + offset_y, f'{idx + 1}', color='red')

    # Display hinges
    hinge_nodes = [0, -2]  # Indices of the two end nodes
    hinge_coordinates = nodes[hinge_nodes]
    plt.plot(hinge_coordinates[:, 0], hinge_coordinates[:, 1], 'r^')  # Display red triangles for hinges

    # Remove axes and labels
    plt.gca().axes.get_xaxis().set_visible(False)
    plt.gca().axes.get_yaxis().set_visible(False)

    # Remove grid
    plt.grid(False)

    plt.title('Truss Structure')

    # Save image to temporary file
    temp_file = 'truss_structure.png'
    plt.savefig(temp_file)
    plt.close()
    return temp_file

#########################################################################

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        load_image('trussback1.png', self.ui.label)
        load_image('6313503-200.png', self.ui.label_24)
        load_image('5469180-200.png', self.ui.label_25)
        load_image('2018888-200.png', self.ui.label_26)

        self.ui.pushButton.clicked.connect(self.on_pushButton_clicked)

    def on_pushButton_clicked(self):
        self.hide()
        self.window2 = MainWindow2()
        self.window2.show()


class MainWindow2(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui2 = Ui_MainWindow2()
        self.ui2.setupUi(self)

        load_image('truss11.jpg', self.ui2.label)
        load_image('truss2.jpg', self.ui2.label_2)
        load_image('truss3.jpg', self.ui2.label_3)
        load_image('truss1.png', self.ui2.label_23)
        load_image('DT2.jpg', self.ui2.label_21)
        load_image('DT.png', self.ui2.label_20)

        self.ui2.pushButton.clicked.connect(self.on_pushButton_clicked)

    def on_pushButton_clicked(self):
        self.hide()
        self.window3 = MainWindow3()
        self.window3.show()


class MainWindow3(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui3 = Ui_MainWindow3()
        self.ui3.setupUi(self)

        load_image('DT3.jpg', self.ui3.label_22)

        self.ui3.pushButton.clicked.connect(self.on_pushButton_clicked)

    def on_pushButton_clicked(self):
        try:
            # 사용자 입력값 가져오기
            material_type = self.ui3.comboBox.currentText()
            bridge_length_m = float(self.ui3.bridgeLengthInput.text().strip())
            support_points_count = int(self.ui3.supportPointsInput.text().strip())
            live_load_kN = float(self.ui3.liveLoadInput.text().strip())
            member_section = float(self.ui3.memberSectionInput.text().strip())

            # 계산 함수 실행, 고정 하중 계산
            self_weight, fixed_load_kN = calculate_weight(member_section, bridge_length_m, support_points_count)

            # 다음 창으로 이동하여 계산 및 결과 표시
            self.hide()
            self.window4 = MainWindow4(material_type, bridge_length_m, support_points_count, live_load_kN,
                                       member_section, fixed_load_kN)
            self.window4.show()


        except ValueError as e:
            print(f"Error: {e}")

###############################################################


class Solver:
    def __init__(self, material_type, bridge_length_m, support_points_count, live_load_kN, member_section,
                 fixed_load_kN):
        self.material_type = material_type
        self.bridge_length_m = bridge_length_m
        self.support_points_count = support_points_count
        self.live_load_kN = live_load_kN
        self.member_section = member_section
        self.fixed_load_kN = fixed_load_kN
        self.material_elasticity_kg_per_mm2 = material_elasticity_kg_per_mm2
        self.ui4 = Ui_MainWindow4()

    def solve(self):
        try:
            # Create BridgeElement instance
            bridge = BridgeElement(self.material_type, self.bridge_length_m, self.support_points_count,
                                   material_elasticity_kg_per_mm2, self.live_load_kN, self.member_section)

            # Set fixed load and unit weight using calculate_weight function
            self_weight, fixed_load_kN = calculate_weight(self.member_section, self.bridge_length_m,
                                                          self.support_points_count)
            bridge.set_fixed_load(fixed_load_kN)
            bridge.set_unit_weight(self_weight)

            # Calculate Mu, D, L
            Mu, D, L = calculate_flexural_strength(bridge.member_section, bridge.bridge_length_m, bridge.live_load_kN,
                                                   bridge.fixed_load_kN)

            # Calculate section modulus
            S = calculate_section_modulus(bridge.member_section)

            # Calculate Mn
            Mn = material_elasticity_kg_per_mm2 * S

            # Evaluate safety
            safety_status = evaluate_safety(Mu, Mn)

            # Truss structure creation and visualization
            n = self.support_points_count // 2
            nodes, elements = create_truss_structure(n)

            # Global stiffness matrix assembly
            K = assemble_global_stiffness(nodes, elements)

            # Boundary conditions and external force definition (example: applying force to the last node)
            fixed_dofs = [0, 1, 2, 3]
            free_dofs = list(set(range(2 * nodes.shape[0])) - set(fixed_dofs))
            F = np.zeros(2 * nodes.shape[0])
            F[4 * (n - 1) + 2] = 1000  # Applying force in the x-direction to the last node

            # Displacement calculation
            K_ff = K[np.ix_(free_dofs, free_dofs)]
            F_f = F[free_dofs]
            U_f = np.linalg.solve(K_ff, F_f)

            # Full displacement vector composition
            U = np.zeros(2 * nodes.shape[0])
            U[free_dofs] = U_f

            # Element stress calculation
            stresses = []
            for element in elements:
                node1 = nodes[element[0]]
                node2 = nodes[element[1]]
                length, angle = element_properties(node1, node2)
                dof_indices = np.array([
                    2 * element[0], 2 * element[0] + 1,
                    2 * element[1], 2 * element[1] + 1
                ])
                u1 = U[dof_indices[:2]]
                u2 = U[dof_indices[2:]]
                stress = element_stress(node1, node2, u1, u2, length, angle)
                stresses.append(stress)

            # Truss structure visualization
            plot_truss(nodes, elements)

            # Prepare solver result text
            solver_result = (
                f"자중으로 인한 등분포하중: {fixed_load_kN:.2f} N/m\n"
                f"자중에 의한 모멘트: {D:.2f} N*m\n"
                f"외부하중에 의한 모멘트: {L:.2f} N*m\n"
                f"계산된 설계휨모멘트 (Mu): {Mu:.2f} N*m\n"
                f"계산된 단면적 모멘트 (S): {S:.6f} m^3\n"
                f"계산된 공칭휨강도 (Mn): {Mn:.2f} N*m\n"
                f"안전성 평가: {safety_status}"
            )

            return solver_result, safety_status

        except ValueError:
            return "계산 오류: 잘못된 입력입니다. 숫자를 입력해주세요."

class MainWindow4(QMainWindow):
    def __init__(self, material_type, bridge_length_m, support_points_count, live_load_kN, member_section, fixed_load_kN):
        super().__init__()
        self.ui4 = Ui_MainWindow4()
        self.ui4.setupUi(self)

        self.material_type = material_type
        self.bridge_length_m = bridge_length_m
        self.support_points_count = support_points_count
        self.live_load_kN = live_load_kN
        self.member_section = member_section
        self.fixed_load_kN = fixed_load_kN

        self.ui4.label_12.setText(f"{self.material_type}")
        self.ui4.label_18.setText(f"{self.bridge_length_m}")
        self.ui4.label_23.setText(f"{self.member_section} m^2")
        self.ui4.label_24.setText(f"{self.support_points_count} 개")
        self.ui4.label_25.setText(f"{self.fixed_load_kN} kN")
        self.ui4.label_26.setText(f"{self.live_load_kN} kN")

        self.ui4.label_31.setText(
            f"재료 종류: {self.material_type}\n"
            f"교량 길이: {self.bridge_length_m} m\n"
            f"절점 갯수: {self.support_points_count}\n"
            f"활 하중: {self.live_load_kN} kN\n"
            f"부재 단면적: {self.member_section} m^2\n"
            f"계산된 고정 하중: {self.fixed_load_kN:.2f} N/m"
        )
        self.ui4.pushButton.clicked.connect(self.save_results_to_excel)
        self.run_solver()

    def run_solver(self):
        try:
            solver = Solver(self.material_type, self.bridge_length_m, self.support_points_count, self.live_load_kN,
                            self.member_section, self.fixed_load_kN)
            self.solver_result, self.safety_status = solver.solve()

            self.ui4.label_27.setText(self.solver_result[:1000])
            self.ui4.label_28.setText(self.safety_status)

            n = self.support_points_count // 2
            self.nodes, self.elements = create_truss_structure(n)

            K = assemble_global_stiffness(self.nodes, self.elements)

            fixed_dofs = [0, 1, 2, 3]
            free_dofs = list(set(range(2 * self.nodes.shape[0])) - set(fixed_dofs))
            F = np.zeros(2 * self.nodes.shape[0])
            F[4 * (n - 1) + 2] = 1000

            K_ff = K[np.ix_(free_dofs, free_dofs)]
            F_f = F[free_dofs]
            U_f = np.linalg.solve(K_ff, F_f)

            U = np.zeros(2 * self.nodes.shape[0])
            U[free_dofs] = U_f

            self.stresses = []
            for element in self.elements:
                node1 = self.nodes[element[0]]
                node2 = self.nodes[element[1]]
                length, angle = element_properties(node1, node2)
                dof_indices = np.array([
                    2 * element[0], 2 * element[0] + 1,
                    2 * element[1], 2 * element[1] + 1
                ])
                u1 = U[dof_indices[:2]]
                u2 = U[dof_indices[2:]]
                stress = element_stress(node1, node2, u1, u2, length, angle)
                self.stresses.append(stress)

            self.temp_file = self.plot_truss(self.nodes, self.elements, self.bridge_length_m)

            image = QImage(self.temp_file)

            pixmap = QPixmap.fromImage(image)
            self.ui4.label.setPixmap(pixmap)
            self.ui4.label.setFixedSize(pixmap.width(), pixmap.height())
            self.ui4.label.setScaledContents(True)

        except ValueError:
            print("계산 오류: 잘못된 입력입니다. 숫자를 입력해주세요.")

    def plot_truss(self, nodes, elements, bridge_length):
        plt.figure(figsize=(8.2, 5))

        for idx, element in enumerate(elements):
            node1 = nodes[element[0]]
            node2 = nodes[element[1]]
            x_values = [node1[0], node2[0]]
            y_values = [node1[1], node2[1]]
            plt.plot(x_values, y_values, 'b-o')

            offset_x = (node2[0] - node1[0]) * 0.05
            offset_y = (node2[1] - node1[1]) * 0.05
            plt.text(np.mean(x_values) + offset_x, np.mean(y_values) + offset_y, f'{idx + 1}', color='red')

        hinge_nodes = [0, -2]
        hinge_coordinates = nodes[hinge_nodes]
        plt.plot(hinge_coordinates[:, 0], hinge_coordinates[:, 1], 'r^')

        plt.gca().axes.get_xaxis().set_visible(False)
        plt.gca().axes.get_yaxis().set_visible(False)
        plt.grid(False)

        plt.title('Truss Structure')

        temp_file = 'truss_structure.png'
        plt.savefig(temp_file)
        plt.close()
        return temp_file

    def save_results_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "트러스 구조 설계 결과"

            title_cell = ws['A1']
            title_cell.value = '트러스 구조 설계 결과'
            title_cell.font = Font(size=25, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:I1')

            ws['A4'] = '날짜'
            ws['B4'] = '2024-06-16'
            ws['B4'].alignment = Alignment(horizontal='right')

            # 이미지 파일 경로를 명시적으로 지정
            image_file_path = 'truss_structure.png'
            if self.temp_file:
                try:
                    img = ExcelImage(image_file_path)
                    img.anchor = 'A6'
                    ws.add_image(img)
                except Exception as img_err:
                    print(f"이미지 파일을 불러오는 중 오류 발생: {img_err}")

            current_row = 30

            result_data = {
                "재료 종류": self.material_type,
                "교량 길이 (m)": self.bridge_length_m,
                "절점 갯수": self.support_points_count,
                "활 하중 (kN)": self.live_load_kN,
                "부재 단면적 (m^2)": self.member_section,
                "계산된 고정 하중 (N/m)": self.fixed_load_kN,
                "안전성 평가": self.safety_status,
            }

            for key, value in result_data.items():
                ws[f'A{current_row}'] = key
                ws.merge_cells(f'B{current_row}:D{current_row}')
                ws[f'B{current_row}'] = str(value)
                current_row += 1

            current_row += 1

            ws[f'A{current_row}'] = "솔버 결과"
            current_row += 1
            solver_lines = self.solver_result.strip().split('\n')
            for line in solver_lines:
                ws[f'A{current_row}'] = line.strip()
                current_row += 1

            current_row += 2

            # 부재별 응력과 길이 데이터
            ws[f'A{current_row}'] = "부재별 응력과 길이"
            current_row += 1
            ws[f'A{current_row}'] = "부재번호"
            ws[f'B{current_row}'] = "응력 (MPa)"
            ws[f'C{current_row}'] = "길이 (m)"

            for idx, (element, stress) in enumerate(zip(self.elements, self.stresses), start=1):
                node1_coords = self.nodes[element[0] - 1]  # 노드1 좌표
                node2_coords = self.nodes[element[1] - 1]  # 노드2 좌표
                length, angle = self.element_properties(node1_coords, node2_coords)
                member_length = length  # 부재의 길이
                current_row += 1
                ws[f'A{current_row}'] = idx
                ws[f'B{current_row}'] = stress
                ws[f'C{current_row}'] = member_length

            current_row += 2

            # 요소 데이터
            ws[f'A{current_row}'] = "요소 데이터"
            elements_df = pd.DataFrame(self.elements, columns=["노드1", "노드2"])
            for r in dataframe_to_rows(elements_df, index=False, header=True):
                current_row += 1
                for col, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=col + 1, value=value)

            current_row += 2

            # 노드 데이터
            ws[f'A{current_row}'] = "노드 데이터"
            nodes_df = pd.DataFrame(self.nodes, columns=["X 좌표", "Y 좌표"])
            for r in dataframe_to_rows(nodes_df, index=False, header=True):
                current_row += 1
                for col, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=col, value=value)

            file_path = '트러스_구조_설계_보고서.xlsx'
            wb.save(file_path)

            print(f"엑셀 파일 저장이 완료되었습니다. 파일 경로: {file_path}")

        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")

    def element_properties(self, node1, node2):
        """Calculate the length and angle of an element."""
        x1, y1 = node1
        x2, y2 = node2
        length = np.linalg.norm(node2 - node1)
        angle = np.arctan2(y2 - y1, x2 - x1)
        return length, angle

    def element_stress(self, node1, node2, u1, u2, E):
        """Calculate the stress of an element."""
        length, angle = self.element_properties(node1, node2)
        c = np.cos(angle)
        s = np.sin(angle)
        T = np.array([-c, -s, c, s])
        strain = (1 / length) * T @ np.concatenate((u1, u2))
        stress = E * strain
        return stress

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())