from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()

ws = wb.active

ws['A1'] = '건설프로그래밍 2조 설계 프로그램 보고서'
ws['A4'] = '날짜'
ws['B4'] = '2024-05-27'
ws['A5'] = '내용'
ws['B6'] = '내용이 들어가는 곳.'

title_font = Font(size=25, bold=True)
ws['A1'].font = title_font

merged_cells_range = 'A1:K3'
for row in ws[merged_cells_range]:
    for cell in row:
        cell.alignment = Alignment(vertical='center',horizontal='center')
ws.merge_cells(merged_cells_range)

date_cell = ws['B4']
date_cell.alignment = Alignment(horizontal='right')

wb.save("건설프로그래밍 2조 보고서.xlsx")